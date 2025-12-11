import base64
import logging
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from config.supabase_client import supabase
from utils.decorators import login_required, superadmin_required

logger = logging.getLogger(__name__)

# Caché simple para contadores
cache_contadores = {
    'data': None,
    'timestamp': None,
    'ttl': 60
}

# Definimos el Blueprint
dashboard_blueprint = Blueprint('dashboard', __name__, template_folder='templates')

# --- HELPER PARA FOTOS (IMÁGENES) ---
def procesar_imagen(file):
    """Convierte la imagen subida a Base64 para guardarla en la BD."""
    if not file or not file.filename:
        return None
    try:
        # Asegurar que estamos al inicio del archivo
        file.seek(0)
        # Leemos los bytes de la imagen
        file_data = file.read()
        # Convertimos a base64
        encoded_image = base64.b64encode(file_data).decode('utf-8')
        # Creamos el string completo (Data URI)
        mime_type = file.content_type or 'image/jpeg'
        return f"data:{mime_type};base64,{encoded_image}"
    except Exception as e:
        print(f"Error procesando imagen: {e}")
        return None

# --- HELPER PARA PDFS (DOCUMENTOS) ---
def procesar_pdf(file):
    """Convierte el PDF a Base64 para guardarlo en la BD."""
    if not file or not file.filename:
        return None
    try:
        file.seek(0)
        file_data = file.read()
        encoded_pdf = base64.b64encode(file_data).decode('utf-8')
        # Forzamos el tipo application/pdf
        return f"data:application/pdf;base64,{encoded_pdf}"
    except Exception as e:
        print(f"Error procesando PDF: {e}")
        return None

# --- HELPER PARA ESTILOS EXCEL ---
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """Función auxiliar para aplicar estilos a un rango de celdas"""
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    
    first_cell = ws[cell_range.split(":")[0]]
    rows = list(ws[cell_range])
    
    for cell in rows[0]: cell.border = cell.border + top
    for cell in rows[-1]: cell.border = cell.border + bottom
    
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        
        for cell in row:
            if font: cell.font = font
            if alignment: cell.alignment = alignment

# --- HELPER PARA RECOLECTAR DATOS DEL FORMULARIO ---
def obtener_datos_formulario(req):
    """Extrae todos los campos del formulario para crear o editar."""
    # Usamos .get() para evitar errores si el campo no existe en el form
    datos = {
        # Datos Básicos
        'nombre': req.form.get('nombre'),
        'apellido': req.form.get('apellido'),
        'cedula': req.form.get('cedula'),
        'edad': req.form.get('edad') or None, # Convertir vacíos a None para enteros
        'sexo': req.form.get('sexo'),
        'email': req.form.get('email'),
        'telefono': req.form.get('telefono'),
        'estatus': req.form.get('estatus'),
        'cuenta_bancaria': req.form.get('cuenta_bancaria'),
        
        # Datos del Representante
        'es_menor': True if req.form.get('es_menor') == 'on' else False,
        'representante_nombre': req.form.get('representante_nombre'),
        'representante_cedula': req.form.get('representante_cedula'),
        'representante_parentesco': req.form.get('representante_parentesco'),
        
        # Ubicación
        'municipio': req.form.get('municipio'),
        'lugar_nacimiento': req.form.get('lugar_nacimiento'),
        'direccion': req.form.get('direccion'),
        'fecha_nacimiento': req.form.get('fecha_nacimiento') or None,

        # Datos Deportivos
        'disciplina': req.form.get('disciplina'),
        'especialidad': req.form.get('especialidad'),
        'categoria': req.form.get('categoria'),
        'tipo_beca': req.form.get('tipo_beca'),

        # Antropometría
        'sangre': req.form.get('sangre'),
        'peso': req.form.get('peso'),
        'estatura': req.form.get('estatura'),

        # Tallas
        'talla_zapato': req.form.get('talla_zapato'),
        'talla_franela': req.form.get('talla_franela'),
        'talla_short': req.form.get('talla_short'),
        'talla_chemise': req.form.get('talla_chemise'),
        'talla_mono': req.form.get('talla_mono'),
        'talla_competencia': req.form.get('talla_competencia'),

        # Información Médica y Otros (Selects Si/No)
        'usa_lentes': req.form.get('usa_lentes'),
        'usa_bucal': req.form.get('usa_bucal'),
        'usa_munequera': req.form.get('usa_munequera'),
        'usa_rodilleras': req.form.get('usa_rodilleras'),
        'dieta_deportiva': req.form.get('dieta_deportiva'),
        'control_medico': req.form.get('control_medico'),
        'estudio_social': req.form.get('estudio_social'),
    }
    return datos

# --- CONTEXT PROCESSOR ---
@dashboard_blueprint.context_processor
def inject_role():
    """Inyecta el rol del usuario en todas las plantillas."""
    return dict(user_role=session.get('role', 'usuario'))

# --- RUTA HOME (INICIO) CON CACHÉ ---
@dashboard_blueprint.route('/')
@login_required
def index():
    """Página de INICIO (Home) con contadores optimizados."""
    try:
        user = supabase.auth.get_user().user
        first_name = user.user_metadata.get('first_name', 'Usuario')

        # Verificar si el caché es válido
        now = datetime.now()
        cache_valido = (
            cache_contadores['data'] is not None and 
            cache_contadores['timestamp'] is not None and
            (now - cache_contadores['timestamp']).total_seconds() < cache_contadores['ttl']
        )
        
        if cache_valido:
            contadores = cache_contadores['data']
        else:
            atletas = supabase.table('becas').select('id', count='exact').eq('estatus', 'Activo').execute().count or 0
            revision = supabase.table('becas').select('id', count='exact').eq('estatus', 'En Revisión').execute().count or 0
            try:
                medallas = supabase.table('medallas').select('id', count='exact').execute().count or 0
            except: 
                medallas = 0
            
            contadores = {'atletas': atletas, 'revision': revision, 'medallas': medallas}
            cache_contadores['data'] = contadores
            cache_contadores['timestamp'] = now
        
        
        
        
        # Cargar imágenes de galería (Slots)
        gallery_images = {}
        try:
            gallery_data = supabase.table('gallery_images').select('*').execute()
            for img in gallery_data.data:
                gallery_images[img['slot']] = img['image_data']
        except:
            pass

        return render_template('dashboard_home.html', 
                               first_name=first_name, 
                               atletas_count=contadores['atletas'], 
                               revision_count=contadores['revision'], 
                               medallas_count=contadores['medallas'],
                               gallery_images=gallery_images)
    except Exception as e:
        logger.error(f"Error cargando dashboard: {e}", exc_info=True)
        session.clear()
        return redirect(url_for('auth.login'))

# --- RUTA LISTADO DE ATLETAS ---
@dashboard_blueprint.route('/becas')
@login_required
def lista_becas():
    """Lista de becas con filtro por disciplina, búsqueda por nombre y paginación."""
    filtro = request.args.get('disciplina')
    busqueda = request.args.get('buscar', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    try:
        start = (page - 1) * per_page
        end = start + per_page - 1
        
        query = supabase.table('becas').select('*', count='exact')
        
        # Filtrar por disciplina
        if filtro and filtro != 'Todas':
            query = query.eq('disciplina', filtro)
        
        # Búsqueda por nombre o apellido (case-insensitive)
        if busqueda:
            query = query.or_(f"nombre.ilike.%{busqueda}%,apellido.ilike.%{busqueda}%")
        
        result = query.order('id', desc=True).range(start, end).execute()
        becas = result.data
        total = result.count or 0
        total_pages = (total + per_page - 1) // per_page if total > 0 else 1
        
        all_d = supabase.table('becas').select('disciplina').limit(1000).execute()
        lista_d = sorted(list(set(i['disciplina'] for i in all_d.data if i.get('disciplina'))))
        
        # Cargar imágenes de galería
        gallery_images = {}
        try:
            gallery_data = supabase.table('gallery_images').select('*').execute()
            for img in gallery_data.data:
                gallery_images[img['slot']] = img['image_data']
        except:
            pass  # Si la tabla no existe, continuar sin imágenes
        
        return render_template('dashboard_becas.html', 
                             becas=becas, 
                             disciplinas=lista_d, 
                             current_filter=filtro,
                             current_search=busqueda,
                             gallery_images=gallery_images,
                             page=page,
                             total_pages=total_pages,
                             total=total)
    except Exception as e:
        logger.error(f"Error al cargar listado: {e}", exc_info=True)
        flash(f'Error al cargar listado: {e}', 'error')
        return render_template('dashboard_becas.html', becas=[], disciplinas=[], page=1, total_pages=1, total=0, current_search='', gallery_images={})

# --- RUTA MI CUENTA ---
@dashboard_blueprint.route('/cuenta', methods=['GET', 'POST'])
@login_required
def mi_cuenta():
    """Gestionar perfil y contraseña."""
    try:
        user = supabase.auth.get_user().user
        if request.method == 'POST':
            # Actualizar Datos Personales
            if 'update_info' in request.form:
                fname = request.form.get('first_name')
                lname = request.form.get('last_name')
                supabase.auth.update_user({"data": {"first_name": fname, "last_name": lname, "full_name": f"{fname} {lname}"}})
                flash('Datos actualizados.', 'success')
            
            # Cambiar Contraseña
            elif 'update_password' in request.form:
                pwd = request.form.get('password')
                if len(pwd) < 6:
                    flash('La contraseña es muy corta.', 'error')
                else:
                    supabase.auth.update_user({"password": pwd})
                    flash('Contraseña actualizada.', 'success')
            return redirect(url_for('dashboard.mi_cuenta'))

        return render_template('dashboard_cuenta.html', first_name=user.user_metadata.get('first_name'), last_name=user.user_metadata.get('last_name'), email=user.email)
    except: return redirect(url_for('dashboard.index'))


# --- GESTIÓN DE ATLETAS (CRUD) ---

@dashboard_blueprint.route('/becas/nueva', methods=['GET', 'POST'])
@login_required
def crear_beca():
    print("="*50)
    print("DEBUG: CREAR_BECA INICIADO")
    print(f"DEBUG: Método: {request.method}")
    
    if request.method == 'POST':
        try:
            print("DEBUG: Procesando POST")
            print(f"DEBUG: request.files: {request.files}")
            print(f"DEBUG: request.form keys: {request.form.keys()}")
            
            # 1. Recolectar todos los datos del formulario usando el helper
            datos = obtener_datos_formulario(request)
            print(f"DEBUG: Datos recolectados: {list(datos.keys())}")
            
            # 2. Procesar Foto
            file = request.files.get('foto')
            print(f"DEBUG CREAR: Archivo recibido: {file}")
            if file:
                print(f"DEBUG CREAR: Filename: {file.filename}")
                print(f"DEBUG CREAR: Content-Type: {file.content_type}")
            
            foto_b64 = procesar_imagen(file)
            print(f"DEBUG CREAR: Resultado procesar_imagen: {foto_b64[:100] if foto_b64 else 'None'}...")
            
            if foto_b64:
                datos['foto'] = foto_b64
                print(f"DEBUG: Foto length: {len(foto_b64)}")
            
            # 3. Insertar en BD
            print("DEBUG: Intentando insertar en BD...")
            result = supabase.table('becas').insert(datos).execute()
            print(f"DEBUG: Insert exitoso, data: {result.data}")
            
            flash('Atleta registrado exitosamente.', 'success')
            return redirect(url_for('dashboard.lista_becas'))
        except Exception as e: 
            print(f"ERROR en crear_beca: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Error al registrar: {e}', 'error')
    return render_template('crear_beca.html')

@dashboard_blueprint.route('/becas/ver/<int:beca_id>')
@login_required
def ver_beca(beca_id):
    """Ver ficha técnica y medallas (Solo lectura)."""
    try:
        # Cargar Atleta
        beca = supabase.table('becas').select('*').eq('id', beca_id).single().execute().data
        
        if not beca:
            flash('Atleta no encontrado.', 'error')
            return redirect(url_for('dashboard.lista_becas'))

        # Cargar Medallas
        try:
            medallas = supabase.table('medallas').select('*').eq('atleta_id', beca_id).order('created_at', desc=True).execute().data
        except: medallas = [] # Si falla (ej. tabla no existe), lista vacía

        # Cargar Documentos (opcional para ver ficha)
        try:
            documentos = supabase.table('documentos').select('*').eq('atleta_id', beca_id).execute().data
        except: documentos = []

        return render_template('ver_beca.html', beca=beca, medallas=medallas, documentos=documentos)
    except Exception as e: 
        flash(f'Error al cargar ficha: {e}', 'error')
        return redirect(url_for('dashboard.lista_becas'))

@dashboard_blueprint.route('/becas/descargar/<int:beca_id>')
@login_required
def descargar_ficha(beca_id):
    try:
        # 1. Obtener datos del atleta
        beca = supabase.table('becas').select('*').eq('id', beca_id).single().execute().data
        if not beca:
            flash('Atleta no encontrado.', 'error')
            return redirect(url_for('dashboard.lista_becas'))

        # 2. Crear el libro y la hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Ficha Atleta"

        # Estilos comunes
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Configurar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15

        # --- ENCABEZADO ---
        ws.merge_cells('A1:H1')
        ws['A1'] = "REPUBLICA BOLIVARIANA DE VENEZUELA"
        ws['A1'].font = bold_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:H2')
        ws['A2'] = "INSTITUTO REGIONAL DE DEPORTES DEL ESTADO BOLIVARIANO DE GUÁRICO"
        ws['A2'].font = bold_font
        ws['A2'].alignment = center_align

        ws.merge_cells('A3:H3')
        ws['A3'] = "PROGRAMA BECAS DEPORTIVAS"
        ws['A3'].font = bold_font
        ws['A3'].alignment = center_align

        # --- DATOS GENERALES ---
        ws['A5'] = "FICHA ATLETA"
        ws['A5'].font = bold_font
        ws['D5'] = "DISCIPLINA DEPORTIVA:"
        ws['D5'].font = bold_font
        ws['G5'] = beca.get('disciplina', '')

        ws['A6'] = "DATOS PERSONALES"
        ws['A6'].font = bold_font
        ws['D6'] = "Especialidad, Modalidad, División:"
        ws['G6'] = beca.get('especialidad', '')
        
        ws['F7'] = "FECHA SOLICITUD:"
        ws['F7'].font = bold_font
        # Usamos created_at o la fecha actual si no hay
        ws['G7'] = beca.get('created_at', '').split('T')[0] if beca.get('created_at') else ''

        # --- TABLA IDENTIFICACIÓN ---
        headers_id = ["NACIONALIDAD", "C. I. N°", "PRIMER APELLIDO", "SEGUNDO APELLIDO", "PRIMER NOMBRE", "SEGUNDO NOMBRE"]
        ws.append([]) # Espacio
        ws.append(headers_id + ["", "FOTO"]) 
        
        # Datos del atleta
        nacionalidad = "Venezolana" 
        cedula = beca.get('cedula', '')
        apellido = beca.get('apellido', '')
        nombre = beca.get('nombre', '')
        
        datos_atleta = [nacionalidad, cedula, apellido, "", nombre, "", "", "[FOTO]"]
        ws.append(datos_atleta)
        
        style_range(ws, 'A9:H10', border=thin_border, alignment=center_align)
        for col in ['A','B','C','D','E','F']:
            ws[f'{col}9'].font = bold_font

        # --- DATOS FISICOS Y NACIMIENTO ---
        ws.append([]) # Row 11 vacía
        ws.append(["FECHA DE NACIMIENTO", "", "", "SEXO", "DIAGNÓSTICO MÉDICO", "", "GRUPO SANGUÍNEO", "PESO (Kg)"]) # Row 12
        ws['A12'].font = bold_font
        
        # Subcabeceras
        ws.append(["DIA", "MES", "AÑO", f"{beca.get('sexo','')} [X]", "No", "", beca.get('sangre',''), beca.get('peso','')]) # Row 13
        ws['I13'] = f"{beca.get('estatura','')} m"
        
        # Valores fecha
        fecha_nac = beca.get('fecha_nacimiento', '')
        dia, mes, anio = "", "", ""
        if fecha_nac:
            try:
                parts = fecha_nac.split('-')
                if len(parts) == 3:
                    anio, mes, dia = parts
            except: pass

        ws.append([dia, mes, anio, "", "", "", "", ""]) # Row 14
        
        style_range(ws, 'A12:H14', border=thin_border, alignment=center_align)

        # --- DIRECCIÓN ---
        ws.append([]) # Row 15 vacía
        ws.merge_cells('A16:F16')
        ws['A16'] = "LUGAR DE NACIMIENTO Y DIRECCIÓN DE HABITACIÓN DEL ATLETA"
        ws['A16'].font = bold_font
        ws['G16'] = "PROGRAMA DEPORTIVO"
        ws['G16'].font = bold_font
        
        ws.append(["PAÍS", "", "ESTADO", "MUNICIPIO", "CIUDAD", "PARROQUIA", "Elite II", ""]) # Row 17
        ws.append(["Venezuela", "", "Guarico", beca.get('municipio',''), "San Juan de los Morros", "", "", ""]) # Row 18
        
        ws.merge_cells('A19:H19')
        ws['A19'] = "DIRECCIÓN: URB - BARRIO - SECTOR - AVENIDA - CALLE - VEREDA - N°CASA"
        ws['A19'].font = bold_font
        
        ws.merge_cells('A20:F20')
        ws['A20'] = beca.get('direccion', '')
        ws['G19'] = "TELÉFONO / EMAIL"
        ws['G20'] = f"{beca.get('telefono','')} / {beca.get('email','')}"

        style_range(ws, 'A16:H20', border=thin_border, alignment=left_align)

        # --- OTROS DATOS (Lentes, tallas, etc) ---
        ws.append([]) # Row 21
        ws.append(["OTRA INFORMACIÓN DEPORTIVA"]) # Row 22
        ws['A22'].font = bold_font
        
        ws.append(["¿USA LENTES?", "¿PROTECTOR BUCAL?", "¿MUÑEQUERA?", "¿RODILLERAS?", "¿DIETA?", "CONTROL MÉDICO"]) # Row 23
        
        def check_si(val): return "Si [X]" if val == 'Si' else "No"
        
        ws.append([
            check_si(beca.get('usa_lentes')),
            check_si(beca.get('usa_bucal')),
            check_si(beca.get('usa_munequera')),
            check_si(beca.get('usa_rodilleras')),
            check_si(beca.get('dieta_deportiva')),
            check_si(beca.get('control_medico'))
        ]) # Row 24
        
        ws.append(["TALLAS:", "Chaqueta/Mono", "Chemise", "Franela", "Short", "Calzado", "Uniforme Comp."]) # Row 25
        ws.append(["", beca.get('talla_mono',''), beca.get('talla_chemise',''), beca.get('talla_franela',''), beca.get('talla_short',''), beca.get('talla_zapato',''), beca.get('talla_competencia','')]) # Row 26
        
        style_range(ws, 'A23:G26', border=thin_border, alignment=center_align)

        # --- OBSERVACIONES Y NOTAS ---
        ws.append([]) # Row 27
        ws['A27'] = "OBSERVACIONES:"
        ws['A27'].font = bold_font
        
        ws.append([]) # Row 28
        ws['A29'] = "NOTA IMPORTANTE - RECAUDOS:"
        ws['A29'].font = bold_font
        
        recaudos = [
            "● Postulación de la Asociación Deportiva del Atleta.",
            "● Imagen legible de la cédula de identidad o Partida de Nacimiento.",
            "● Síntesis Curricular o Palmarés deportivo.",
            "● Copia de cédula del titular de la cuenta bancaria.",
            "● Copia de cédula del Representante (si aplica).",
            "● Soporte de Cuenta Bancaria (Referencia, Cheque, Libreta)."
        ]
        
        current_row = 30
        for recaudo in recaudos:
            ws.cell(row=current_row, column=1, value=recaudo)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            current_row += 1

        # 3. Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Ficha_{beca.get('nombre','').replace(' ','_')}_{beca.get('apellido','').replace(' ','_')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Error al generar ficha: {e}', 'error')
        return redirect(url_for('dashboard.ver_beca', beca_id=beca_id))

@dashboard_blueprint.route('/becas/editar/<int:beca_id>', methods=['GET', 'POST'])
@login_required
def editar_beca(beca_id):
    """Editar datos y gestionar medallas."""
    
    # 1. Guardar cambios del atleta (POST)
    if request.method == 'POST':
        try:
            # Recolectar datos actualizados
            datos = obtener_datos_formulario(request)
            
            # Procesar Foto Nueva (si se subió una)
            file = request.files.get('foto')
            print(f"DEBUG: Archivo recibido: {file}")
            if file:
                print(f"DEBUG: Filename: {file.filename}")
            
            foto_b64 = procesar_imagen(file)
            print(f"DEBUG: Resultado procesar_imagen: {foto_b64[:50] if foto_b64 else 'None'}")
            
            if foto_b64:
                datos['foto'] = foto_b64
                print("DEBUG: Foto agregada a datos para update")
            
            supabase.table('becas').update(datos).eq('id', beca_id).execute()
            flash('Ficha actualizada correctamente.', 'success')
            return redirect(url_for('dashboard.editar_beca', beca_id=beca_id))
        except Exception as e: 
            flash(f'Error al guardar cambios: {e}', 'error')

    # 2. MOSTRAR (GET)
    try:
        beca = supabase.table('becas').select('*').eq('id', beca_id).single().execute().data
        
        # Cargar medallas
        try:
            medallas = supabase.table('medallas').select('*').eq('atleta_id', beca_id).order('created_at', desc=True).execute().data
        except: medallas = []

        # Cargar documentos
        try:
            documentos = supabase.table('documentos').select('*').eq('atleta_id', beca_id).order('created_at', desc=True).execute().data
        except: documentos = []
        
        return render_template('editar_beca.html', beca=beca, medallas=medallas, documentos=documentos)
    except: return redirect(url_for('dashboard.lista_becas'))

@dashboard_blueprint.route('/becas/eliminar/<int:beca_id>', methods=['POST'])
@login_required
def eliminar_beca(beca_id):
    try:
        supabase.table('becas').delete().eq('id', beca_id).execute()
        flash('Atleta eliminado.', 'success')
    except Exception as e: flash(f'Error: {e}', 'error')
    return redirect(url_for('dashboard.lista_becas'))

# --- GESTIÓN DE MEDALLAS (Rutas Auxiliares) ---

@dashboard_blueprint.route('/becas/<int:beca_id>/agregar_medalla', methods=['POST'])
@login_required
def agregar_medalla(beca_id):
    try:
        supabase.table('medallas').insert({
            'atleta_id': beca_id,
            'tipo_medalla': request.form.get('tipo_medalla'),
            'competicion': request.form.get('competicion'),
            'fecha': request.form.get('fecha') or None
        }).execute()
        flash('Medalla agregada.', 'success')
    except Exception as e: flash(f'Error al agregar medalla: {e}', 'error')
    # Volver a la misma página de edición
    return redirect(url_for('dashboard.editar_beca', beca_id=beca_id))

@dashboard_blueprint.route('/medallas/eliminar/<int:medalla_id>', methods=['POST'])
@login_required
def eliminar_medalla(medalla_id):
    atleta_id = request.form.get('atleta_id') # Necesario para saber a dónde volver
    try:
        supabase.table('medallas').delete().eq('id', medalla_id).execute()
        flash('Medalla eliminada.', 'success')
    except: flash('Error al eliminar.', 'error')
    return redirect(url_for('dashboard.editar_beca', beca_id=atleta_id))

# --- GESTIÓN DE DOCUMENTOS (Rutas Auxiliares) ---

@dashboard_blueprint.route('/becas/<int:beca_id>/subir_documento', methods=['POST'])
@login_required
def subir_documento(beca_id):
    try:
        archivo = request.files.get('archivo_pdf')
        nombre_doc = request.form.get('nombre_doc')
        
        if not archivo or not nombre_doc:
            flash('Debes seleccionar un archivo y un nombre.', 'error')
        else:
            pdf_b64 = procesar_pdf(archivo)
            if pdf_b64:
                supabase.table('documentos').insert({
                    'atleta_id': beca_id,
                    'nombre': nombre_doc,
                    'archivo': pdf_b64
                }).execute()
                flash('Documento subido correctamente.', 'success')
            else:
                flash('Error al procesar el archivo PDF.', 'error')
    except Exception as e: 
        flash(f'Error al subir: {e}', 'error')
    
    return redirect(url_for('dashboard.editar_beca', beca_id=beca_id))

@dashboard_blueprint.route('/documentos/eliminar/<int:doc_id>', methods=['POST'])
@login_required
def eliminar_documento(doc_id):
    atleta_id = request.form.get('atleta_id')
    try:
        supabase.table('documentos').delete().eq('id', doc_id).execute()
        flash('Documento eliminado.', 'success')
    except: flash('Error al eliminar.', 'error')
    return redirect(url_for('dashboard.editar_beca', beca_id=atleta_id))


# --- GESTIÓN DE USUARIOS (SOLO SUPERADMIN) ---

@dashboard_blueprint.route('/usuarios')
@login_required
@superadmin_required
def lista_usuarios():
    try:
        usuarios = supabase.table('profiles').select('*').order('email').execute().data
        return render_template('dashboard_usuarios.html', usuarios=usuarios)
    except Exception as e:
        logger.error(f"Error cargando usuarios: {e}", exc_info=True)
        flash(f'Error al cargar usuarios: {e}', 'error')
        return redirect(url_for('dashboard.index'))

@dashboard_blueprint.route('/usuarios/cambiar_rol', methods=['POST'])
@login_required
@superadmin_required
def cambiar_rol():
    uid = request.form.get('user_id')
    role = request.form.get('new_role')
    try:
        if uid == session.get('user_id') and role != 'superadmin':
            flash('No puedes quitarte el rol de Superadmin a ti mismo.', 'error')
        else:
            supabase.table('profiles').update({'role': role}).eq('id', uid).execute()
            flash('Rol actualizado.', 'success')
    except Exception as e:
        logger.error(f"Error cambiando rol: {e}")
        flash('Error al cambiar rol.', 'error')
    return redirect(url_for('dashboard.lista_usuarios'))

@dashboard_blueprint.route('/usuarios/eliminar', methods=['POST'])
@login_required
@superadmin_required
def eliminar_usuario():
    uid = request.form.get('user_id')
    try:
        if uid == session.get('user_id'):
            flash('No puedes eliminar tu propia cuenta.', 'error')
        else:
            # Nota: Esto solo elimina el perfil. Para eliminar el usuario de Auth se requiere Service Key y:
            # supabase.auth.admin.delete_user(uid)
            supabase.table('profiles').delete().eq('id', uid).execute()
            flash('Usuario eliminado (Perfil).', 'success')
    except Exception as e:
        logger.error(f"Error eliminando usuario: {e}")
        flash(f'Error al eliminar usuario: {e}', 'error')
    return redirect(url_for('dashboard.lista_usuarios'))


# --- GESTIÓN DE GALERÍA DE IMÁGENES ---

@dashboard_blueprint.route('/gallery/upload', methods=['POST'])
@login_required
def upload_gallery_image():
    """Subir imagen para la galería del dashboard (Solo admins)."""
    from flask import jsonify
    
    # Verificar permisos
    role = session.get('role', 'usuario')
    if role not in ['admin', 'superadmin']:
        return jsonify({'error': 'No autorizado'}), 403
    
    try:
        file = request.files.get('image')
        slot = request.form.get('slot')
        
        if not file or not slot:
            return jsonify({'error': 'Faltan datos'}), 400
        
        # Validar slot válido (becas laterales fijos, home dinámico)
        if not (slot.startswith('home_') or slot in ['left_1', 'left_2', 'left_3', 'right_1', 'right_2', 'right_3']):
            return jsonify({'error': 'Slot inválido'}), 400
        
        try:
            import base64
            from io import BytesIO
            from PIL import Image
            
            # Procesar imagen a Base64
            file.seek(0)
            img = Image.open(file)
            
            # Convertir a RGB si es necesario
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            # Redimensionar si es muy grande (opcional pero recomendado para base64)
            img.thumbnail((800, 800))
            
            buffer = BytesIO()
            img.save(buffer, format='PNG', quality=85)
            img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
            img_data = f"data:image/png;base64,{img_str}"
            
            # Upsert en la tabla gallery_images
            data = {
                'slot': slot,
                'image_data': img_data,
                'updated_at': 'now()'
            }
            supabase.table('gallery_images').upsert(data).execute()
            
            logger.info(f"Imagen actualizada en tabla gallery_images: {slot}")
            return jsonify({'success': True, 'message': 'Imagen actualizada'}), 200
            
        except Exception as save_error:
            logger.error(f"Error procesando/guardando imagen: {save_error}")
            return jsonify({'error': f'Error procesando imagen: {str(save_error)}'}), 500
        
    except Exception as e:
        logger.error(f"Error subiendo imagen galería: {e}")
        return jsonify({'error': str(e)}), 500