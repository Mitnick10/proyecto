from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """Función auxiliar para aplicar estilos a un rango de celdas"""
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    
    first_cell = ws[cell_range.split(":")[0]]
    rows = list(ws[cell_range])
    
    for cell in rows[0]: cell.border = cell.border + top
    for cell in rows[-1]: cell.border = cell.border + bottom
    
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        
        for cell in row:
            if font: cell.font = font
            if alignment: cell.alignment = alignment

def generar_ficha_excel(beca):
    """
    Genera el archivo Excel de la ficha del atleta.
    Retorna un objeto BytesIO con el contenido del archivo excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Atleta"

    # Estilos comunes
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Configurar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    # --- ENCABEZADO ---
    ws.merge_cells('A1:H1')
    ws['A1'] = "REPUBLICA BOLIVARIANA DE VENEZUELA"
    ws['A1'].font = bold_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:H2')
    ws['A2'] = "INSTITUTO REGIONAL DE DEPORTES DEL ESTADO BOLIVARIANO DE GUÁRICO"
    ws['A2'].font = bold_font
    ws['A2'].alignment = center_align

    ws.merge_cells('A3:H3')
    ws['A3'] = "PROGRAMA BECAS DEPORTIVAS"
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align

    # --- DATOS GENERALES ---
    ws['A5'] = "FICHA ATLETA"
    ws['A5'].font = bold_font
    ws['D5'] = "DISCIPLINA DEPORTIVA:"
    ws['D5'].font = bold_font
    ws['G5'] = beca.get('disciplina', '')

    ws['A6'] = "DATOS PERSONALES"
    ws['A6'].font = bold_font
    ws['D6'] = "Especialidad, Modalidad, División:"
    ws['G6'] = beca.get('especialidad', '')
    
    ws['F7'] = "FECHA SOLICITUD:"
    ws['F7'].font = bold_font
    # Usamos created_at o la fecha actual si no hay
    ws['G7'] = beca.get('created_at', '').split('T')[0] if beca.get('created_at') else ''

    # --- TABLA IDENTIFICACIÓN ---
    headers_id = ["NACIONALIDAD", "C. I. N°", "PRIMER APELLIDO", "SEGUNDO APELLIDO", "PRIMER NOMBRE", "SEGUNDO NOMBRE"]
    ws.append([]) # Espacio
    ws.append(headers_id + ["", "FOTO"]) 
    
    # Datos del atleta
    nacionalidad = "Venezolana" 
    cedula = beca.get('cedula', '')
    apellido = beca.get('apellido', '')
    nombre = beca.get('nombre', '')
    
    datos_atleta = [nacionalidad, cedula, apellido, "", nombre, "", "", "[FOTO]"]
    ws.append(datos_atleta)
    
    style_range(ws, 'A9:H10', border=thin_border, alignment=center_align)
    for col in ['A','B','C','D','E','F']:
        ws[f'{col}9'].font = bold_font

    # --- DATOS FISICOS Y NACIMIENTO ---
    ws.append([]) # Row 11 vacía
    ws.append(["FECHA DE NACIMIENTO", "", "", "SEXO", "DIAGNÓSTICO MÉDICO", "", "GRUPO SANGUÍNEO", "PESO (Kg)"]) # Row 12
    ws['A12'].font = bold_font
    
    # Subcabeceras
    ws.append(["DIA", "MES", "AÑO", f"{beca.get('sexo','')} [X]", "No", "", beca.get('sangre',''), beca.get('peso','')]) # Row 13
    ws['I13'] = f"{beca.get('estatura','')} m"
    
    # Valores fecha
    fecha_nac = beca.get('fecha_nacimiento', '')
    dia, mes, anio = "", "", ""
    if fecha_nac:
        try:
            parts = fecha_nac.split('-')
            if len(parts) == 3:
                anio, mes, dia = parts
        except: pass

    ws.append([dia, mes, anio, "", "", "", "", ""]) # Row 14
    
    style_range(ws, 'A12:H14', border=thin_border, alignment=center_align)

    # --- DIRECCIÓN ---
    ws.append([]) # Row 15 vacía
    ws.merge_cells('A16:F16')
    ws['A16'] = "LUGAR DE NACIMIENTO Y DIRECCIÓN DE HABITACIÓN DEL ATLETA"
    ws['A16'].font = bold_font
    ws['G16'] = "PROGRAMA DEPORTIVO"
    ws['G16'].font = bold_font
    
    ws.append(["PAÍS", "", "ESTADO", "MUNICIPIO", "CIUDAD", "PARROQUIA", "Elite II", ""]) # Row 17
    ws.append(["Venezuela", "", "Guarico", beca.get('municipio',''), "San Juan de los Morros", "", "", ""]) # Row 18
    
    ws.merge_cells('A19:H19')
    ws['A19'] = "DIRECCIÓN: URB - BARRIO - SECTOR - AVENIDA - CALLE - VEREDA - N°CASA"
    ws['A19'].font = bold_font
    
    ws.merge_cells('A20:F20')
    ws['A20'] = beca.get('direccion', '')
    ws['G19'] = "TELÉFONO / EMAIL"
    ws['G20'] = f"{beca.get('telefono','')} / {beca.get('email','')}"

    style_range(ws, 'A16:H20', border=thin_border, alignment=left_align)

    # --- OTROS DATOS (Lentes, tallas, etc) ---
    ws.append([]) # Row 21
    ws.append(["OTRA INFORMACIÓN DEPORTIVA"]) # Row 22
    ws['A22'].font = bold_font
    
    ws.append(["¿USA LENTES?", "¿PROTECTOR BUCAL?", "¿MUÑEQUERA?", "¿RODILLERAS?", "¿DIETA?", "CONTROL MÉDICO"]) # Row 23
    
    def check_si(val): return "Si [X]" if val == 'Si' else "No"
    
    ws.append([
        check_si(beca.get('usa_lentes')),
        check_si(beca.get('usa_bucal')),
        check_si(beca.get('usa_munequera')),
        check_si(beca.get('usa_rodilleras')),
        check_si(beca.get('dieta_deportiva')),
        check_si(beca.get('control_medico'))
    ]) # Row 24
    
    ws.append(["TALLAS:", "Chaqueta/Mono", "Chemise", "Franela", "Short", "Calzado", "Uniforme Comp."]) # Row 25
    ws.append(["", beca.get('talla_mono',''), beca.get('talla_chemise',''), beca.get('talla_franela',''), beca.get('talla_short',''), beca.get('talla_zapato',''), beca.get('talla_competencia','')]) # Row 26
    
    style_range(ws, 'A23:G26', border=thin_border, alignment=center_align)

    # --- OBSERVACIONES Y NOTAS ---
    ws.append([]) # Row 27
    ws['A27'] = "OBSERVACIONES:"
    ws['A27'].font = bold_font
    
    ws.append([]) # Row 28
    ws['A29'] = "NOTA IMPORTANTE - RECAUDOS:"
    ws['A29'].font = bold_font
    
    recaudos = [
        "● Postulación de la Asociación Deportiva del Atleta.",
        "● Imagen legible de la cédula de identidad o Partida de Nacimiento.",
        "● Síntesis Curricular o Palmarés deportivo.",
        "● Copia de cédula del titular de la cuenta bancaria.",
        "● Copia de cédula del Representante (si aplica).",
        "● Soporte de Cuenta Bancaria (Referencia, Cheque, Libreta)."
    ]
    
    current_row = 30
    for recaudo in recaudos:
        ws.cell(row=current_row, column=1, value=recaudo)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        current_row += 1

    # 3. Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
